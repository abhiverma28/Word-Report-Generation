import os
from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import load_workbook
from docx import Document

app = Flask(__name__)
app.secret_key = "d83fc5b8be09be13b6388b1c39be6a38"  # Replace with a secret key for flashing messages

# Ensure that the "temp" directory exists
if not os.path.exists("temp"):
    os.makedirs("temp")

def generate_report(excel_file, word_template_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    # Get all rows with data from the Excel file
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not data_rows:
        flash("No data found in the Excel file.", "error")
        return

    report_count = 0  # To keep track of the number of reports generated
    for data_row in data_rows:
        word_template = Document(word_template_file)  # Create a new template for each row of data

        data_dict = {}
        for idx, value in enumerate(data_row, start=1):
            column_name = ws.cell(row=1, column=idx).value
            if column_name is None:
                column_name = ""
            if column_name not in data_dict:
                data_dict[column_name] = []
            data_dict[column_name].append(str(value))

        # Get the filename from the first column (column A)
        filename = data_row[0]
        if not filename:
            flash("Filename not available for a row. Skipping.", "error")
            continue

        for paragraph in word_template.paragraphs:
            for key in data_dict:
                placeholder = "{{" + key + "}}"
                if placeholder in paragraph.text:
                    paragraph.text = paragraph.text.replace(placeholder, ", ".join(data_dict[key]))

        # Get the directory path of the Excel file
        excel_dir_path = os.path.dirname(excel_file)

        # Combine the directory path and filename to get the full output path
        output_file = os.path.join(excel_dir_path, filename + ".docx")

        with open(output_file, "wb") as f:
            word_template.save(f)
        report_count += 1

    flash(f"{len(data_rows)} data rows were found, and {report_count} reports have been generated.", "success")

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        excel_file = request.files["excelFile"]
        word_template_file = request.files["wordTemplate"]

        if excel_file and word_template_file:
            # Save the uploaded files to temporary locations
            excel_file_path = os.path.join("temp", excel_file.filename)
            word_template_file_path = os.path.join("temp", word_template_file.filename)
            excel_file.save(excel_file_path)
            word_template_file.save(word_template_file_path)

            generate_report(excel_file_path, word_template_file_path)

            # Delete the temporary files after generating the reports
            os.remove(excel_file_path)
            os.remove(word_template_file_path)

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
